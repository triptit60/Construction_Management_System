from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from django.db.models import Sum
from django.contrib.auth.models import User

from projects.models import Project
from inventory.models import InventoryItem
from finance.models import Expense
from inventory.models import Supplier
from django.db.models import F
from django.db import models

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

#pdf export
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from django.utils import timezone

def reports_dashboard(request):
    total_projects = Project.objects.count()
    total_inventory = InventoryItem.objects.count()
    total_inventory_value = sum(item.total_value for item in InventoryItem.objects.all())
    total_expenses = ( Expense.objects.aggregate(total=Sum("amount"))["total"] or 0)
    low_stock = sum(1 for item in InventoryItem.objects.all()if item.is_low_stock)
    context = {
        "total_projects": total_projects,
        "total_inventory": total_inventory,
        "total_inventory_value": total_inventory_value,
        "total_expenses": total_expenses,
        "low_stock": low_stock,
    }
    return render(request, "reports/dashboard.html", context)


def inventory_report(request):
    inventory = InventoryItem.objects.select_related('supplier', 'project')

            # Get filter values
    project = request.GET.get('project')
    supplier = request.GET.get('supplier')
    search = request.GET.get('search')
    low_stock = request.GET.get('low_stock')

    # Apply filters
    if project:
        inventory = inventory.filter(project_id=project)

    if supplier:
        inventory = inventory.filter(supplier_id=supplier)

    if search:
        inventory = inventory.filter(name__icontains=search)

    if low_stock:
        inventory = inventory.filter(
            quantity__lte=models.F('low_stock_threshold')
        )

    context = { 'inventory': inventory, 'projects': Project.objects.all(),'suppliers': Supplier.objects.all(),}
    return render(request, 'reports/inventory_report.html', context)



def expense_report(request):
    expenses = Expense.objects.select_related('project','logged_by').order_by('-date_incurred')

             
    # Get filter values
    project = request.GET.get('project')
    category = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Apply filters
    if project:
        expenses = expenses.filter(project_id=project)

    if category:
        expenses = expenses.filter(category=category)

    if date_from:
        expenses = expenses.filter(date_incurred__gte=date_from)

    if date_to:
        expenses = expenses.filter(date_incurred__lte=date_to)


    context = {'expenses': expenses,'projects': Project.objects.all(),
        'categories': Expense.CATEGORY_CHOICES,}
    return render(request, 'reports/expense_report.html', context)



def project_report(request):
    projects = Project.objects.select_related('manager','created_by').all()

          # Get filter values
    status = request.GET.get('status')
    manager = request.GET.get('manager')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply filters
    if status:
        projects = projects.filter(status=status)

    if manager:
        projects = projects.filter(manager_id=manager)

    if start_date:
        projects = projects.filter(start_date__gte=start_date)

    if end_date:
        projects = projects.filter(end_date__lte=end_date)


    context = {'projects': projects,'managers': User.objects.all(),
        'statuses': Project.STATUS_CHOICES,}
    return render(request, 'reports/project_report.html', context)


def export_project_excel(request):
    projects = Project.objects.select_related("manager", "created_by").all()

    # Apply the same filters
    status = request.GET.get("status")
    manager = request.GET.get("manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if status:
        projects = projects.filter(status=status)

    if manager:
        projects = projects.filter(manager_id=manager)

    if start_date:
        projects = projects.filter(start_date__gte=start_date)

    if end_date:
        projects = projects.filter(end_date__lte=end_date)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Projects Report"

    headers = [
        "Project Code",
        "Project Name",
        "Client",
        "Manager",
        "Status",
        "Budget",
        "Start Date",
        "End Date",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, project in enumerate(projects, 2):
        worksheet.cell(row=row_num, column=1).value = project.project_code
        worksheet.cell(row=row_num, column=2).value = project.name
        worksheet.cell(row=row_num, column=3).value = project.client
        worksheet.cell(row=row_num, column=4).value = (
            project.manager.username if project.manager else ""
        )
        worksheet.cell(row=row_num, column=5).value = project.get_status_display()
        worksheet.cell(row=row_num, column=6).value = float(project.budget)
        worksheet.cell(row=row_num, column=7).value = (
            project.start_date.strftime("%Y-%m-%d")
            if project.start_date else ""
        )
        worksheet.cell(row=row_num, column=8).value = (
            project.end_date.strftime("%Y-%m-%d")
            if project.end_date else ""
        )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = ('attachment; filename="project_report.xlsx"')

    workbook.save(response)
    return response



def export_expense_excel(request):
    expenses = Expense.objects.select_related(
        "project", "logged_by"
    ).order_by("-date_incurred")

    project = request.GET.get("project")
    category = request.GET.get("category")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if project:
        expenses = expenses.filter(project_id=project)

    if category:
        expenses = expenses.filter(category=category)

    if date_from:
        expenses = expenses.filter(date_incurred__gte=date_from)

    if date_to:
        expenses = expenses.filter(date_incurred__lte=date_to)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Expenses Report"

    headers = [
        "Date Incurred",
        "Project",
        "Category",
        "Amount",
        "Description",
        "Logged By",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, expense in enumerate(expenses, 2):
        worksheet.cell(row=row_num, column=1).value = (
            expense.date_incurred.strftime("%Y-%m-%d")
            if expense.date_incurred else ""
        )
        worksheet.cell(row=row_num, column=2).value = (
            expense.project.name if expense.project else ""
        )
        worksheet.cell(row=row_num, column=3).value = expense.get_category_display()
        worksheet.cell(row=row_num, column=4).value = float(expense.amount)
        worksheet.cell(row=row_num, column=5).value = expense.description
        worksheet.cell(row=row_num, column=6).value = (
            expense.logged_by.username if expense.logged_by else ""
        )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="expense_report.xlsx"'
    workbook.save(response)
    return response 


def export_inventory_excel(request):
    inventory = InventoryItem.objects.select_related("supplier", "project")

    project = request.GET.get("project")
    supplier = request.GET.get("supplier")
    search = request.GET.get("search")
    low_stock = request.GET.get("low_stock")

    if project:
        inventory = inventory.filter(project_id=project)

    if supplier:
        inventory = inventory.filter(supplier_id=supplier)

    if search:
        inventory = inventory.filter(name__icontains=search)

    if low_stock:
        inventory = inventory.filter(quantity__lte=F("low_stock_threshold"))

 # Create workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Report"

    headers = [
        "Name",
        "Project",
        "Supplier",
        "Quantity",
        "Low Stock Threshold",
        "Total Value",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, item in enumerate(inventory, 2):
        worksheet.cell(row=row_num, column=1).value = item.name
        worksheet.cell(row=row_num, column=2).value = item.project.name if item.project else ""
        worksheet.cell(row=row_num, column=3).value = item.supplier.name if item.supplier else ""
        worksheet.cell(row=row_num, column=4).value = item.quantity
        worksheet.cell(row=row_num, column=5).value = item.low_stock_threshold
        worksheet.cell(row=row_num, column=6).value = float(item.total_value)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="inventory_report.xlsx"'
    workbook.save(response)
    return response



def export_project_pdf(request):
    projects = Project.objects.select_related(
        "manager", "created_by"
    ).all()

    # Apply the same filters
    status = request.GET.get("status")
    manager = request.GET.get("manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if status:
        projects = projects.filter(status=status)

    if manager:
        projects = projects.filter(manager_id=manager)

    if start_date:
        projects = projects.filter(start_date__gte=start_date)

    if end_date:
        projects = projects.filter(end_date__lte=end_date)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="project_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter)
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("<b>BuildPro ERP - Project Report</b>", styles["Title"])
    elements.append(title)

    generated = Paragraph(
        f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    )
    elements.append(generated)

    elements.append(Paragraph("<br/>", styles["Normal"]))

    data = [[
        "Code",
        "Project",
        "Client",
        "Manager",
        "Status",
        "Budget",
        "Start",
        "End",
    ]]

    for project in projects:
        data.append([
            project.project_code,
            project.name,
            project.client,
            project.manager.username if project.manager else "",
            project.get_status_display(),
            f"Rs. {project.budget}",
            str(project.start_date),
            str(project.end_date),
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 1, colors.black),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
    ]))

    elements.append(table)

    doc.build(elements)

    return response



def export_expense_pdf(request):
    expenses = Expense.objects.select_related(
        "project", "logged_by"
    ).order_by("-date_incurred")

    # Apply filters
    project = request.GET.get("project")
    category = request.GET.get("category")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if project:
        expenses = expenses.filter(project_id=project)

    if category:
        expenses = expenses.filter(category=category)

    if date_from:
        expenses = expenses.filter(date_incurred__gte=date_from)

    if date_to:
        expenses = expenses.filter(date_incurred__lte=date_to)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="expense_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter)
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        "<b>BuildPro ERP - Expense Report</b>",
        styles["Title"]
    )
    elements.append(title)

    generated = Paragraph(
        f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    )
    elements.append(generated)

    elements.append(Paragraph("<br/>", styles["Normal"]))

    data = [[
        "Date",
        "Project",
        "Category",
        "Amount",
        "Description",
        "Logged By",
    ]]

    for expense in expenses:
        data.append([
            str(expense.date_incurred),
            expense.project.name if expense.project else "",
            expense.get_category_display(),
            f"Rs. {expense.amount}",
            expense.description,
            expense.logged_by.username if expense.logged_by else "",
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 1, colors.black),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
    ]))

    elements.append(table)

    doc.build(elements)

    return response




def export_inventory_pdf(request):
    inventory = InventoryItem.objects.select_related(
        "supplier", "project"
    )

    # Apply filters
    project = request.GET.get("project")
    supplier = request.GET.get("supplier")
    search = request.GET.get("search")
    low_stock = request.GET.get("low_stock")

    if project:
        inventory = inventory.filter(project_id=project)

    if supplier:
        inventory = inventory.filter(supplier_id=supplier)

    if search:
        inventory = inventory.filter(name__icontains=search)

    if low_stock:
        inventory = inventory.filter(
            quantity__lte=F("low_stock_threshold")
        )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="inventory_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter)
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        "<b>BuildPro ERP - Inventory Report</b>",
        styles["Title"]
    )
    elements.append(title)

    generated = Paragraph(
        f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    )
    elements.append(generated)

    elements.append(Paragraph("<br/>", styles["Normal"]))

    data = [[
        "Material",
        "Project",
        "Supplier",
        "Quantity",
        "Unit",
        "Unit Price",
        "Total Value",
        "Status",
    ]]

    for item in inventory:
        data.append([
            item.name,
            item.project.name if item.project else "",
            item.supplier.name if item.supplier else "",
            item.quantity,
            item.unit,
            f"Rs. {item.unit_price}",
            f"Rs. {item.total_value}",
            "Low Stock" if item.is_low_stock else "In Stock",
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 1, colors.black),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
    ]))

    elements.append(table)

    doc.build(elements)

    return response